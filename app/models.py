from datetime import datetime

import arrow
import openpyxl
from flask import url_for
from openpyxl import worksheet
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from typing import Any, Dict, List, Optional

from app import db
from app.exceptions import ValidationError


class Project(db.Model):
    __tablename__ = 'projects'
    pid: int = db.Column(db.Integer, primary_key=True)
    reference_number: str = db.Column(db.String, nullable=False)
    name: str = db.Column(db.String, nullable=False)
    margin: float = db.Column(db.Float, nullable=False)
    active: bool = db.Column(db.Boolean, default=True)
    admin_fee: float = db.Column(db.Float, nullable=True)

    client: Optional[Client] = db.relation('Client', uselist=False, backref='project', cascade="all, delete-orphan",
                                           lazy='select')
    superintendent: Optional[Superintendent] = db.relation('Superintendent', uselist=False, backref='project',
                                                           cascade="all, delete-orphan", lazy='select')
    variations: List[Variation] = db.relationship('Variation', backref='project', cascade="all, delete-orphan",
                                                  lazy='select')
    progress_items: List[ProgressItem] = db.relation('ProgressItem', backref='project', cascade="all, delete-orphan",
                                                     lazy='select')

    def __repr__(self) -> str:
        return f'<Project Name: {self.name}, Active: {self.active}, Margin: {self.margin}>'

    def to_json(self) -> Dict[str, Any]:
        if self.superintendent is None:
            superintendent_name = ''
        else:
            superintendent_name = self.superintendent.name

        json_project = {
            'id': self.pid,
            'name': self.name,
            'reference_number': self.reference_number,
            'active': self.active,
            'margin': self.margin,
            'admin_fee': self.admin_fee,
            'variations': url_for('api.get_project_variations', project_id=self.pid, _external=True),
            'progress_items': url_for('api.get_project_progress_items', project_id=self.pid, _external=True),
            'superintendent_name': superintendent_name,
            'url': url_for('api.get_project', project_id=self.pid, _external=True),
        }
        return json_project

    @staticmethod
    def from_json(json_project: Dict[str, Any]) -> 'Project':
        name = json_project.get('name')
        reference_number = json_project.get('reference_number')
        margin = json_project.get('margin')
        admin_fee = json_project.get('admin_fee')
        if name is None or name == '':
            raise ValidationError('project does not have a name')
        return Project(name=name, margin=margin, admin_fee=admin_fee, reference_number=reference_number)

    def export(self) -> str:
        def cm_to_inch(cm: float) -> float:
            return cm * 0.393701

        def prepare(ws: openpyxl.worksheet.Worksheet) -> None:
            ws.header_footer.setHeader(
                    '&L&"Calibri,Regular"&K000000&G&C&"Lao UI,Bold"&8Total Project Construction Pty. Ltd.&"Lao UI,Regular"&K000000_x000D_ACN 117 578 560  ABN 84 117 578 560_x000D_PO Box 313 HALL ACT_x000D_P: 02-6230 2455   F:02-6230 2488_x000D_E: troy@totalproject.com.au')
            ws.header_footer.setFooter(
                    f'&L&"Arial,Italic"&9&K000000App. A - Contract Variations&R&"Arial,Italic"&9&K000000{self.name}')
            ws.page_margins.top = cm_to_inch(3.4)
            ws.page_margins.bottom = cm_to_inch(2)
            ws.page_margins.left = cm_to_inch(1.2)
            ws.page_margins.right = cm_to_inch(1.1)

        fill = PatternFill(patternType='solid', fgColor=Color('D8E4BC'))

        wb: openpyxl.Workbook = openpyxl.Workbook()

        ws: openpyxl.worksheet.Worksheet = wb.active
        ws.title = 'Claim - TOTAL'
        prepare(ws)

        if self.client is None:
            client_name = ''
            client_first_line_address = ''
            client_second_line_address = ''
        else:
            client_name = self.client.name
            client_first_line_address = self.client.first_line_address
            client_second_line_address = self.client.second_line_address

        if self.superintendent is None:
            superintendent_name = ''
            superintendent_first_line_address = ''
            superintendent_second_line_address = ''
        else:
            superintendent_name = self.superintendent.name
            superintendent_first_line_address = self.superintendent.first_line_address
            superintendent_second_line_address = self.superintendent.second_line_address

        ws['A1'].value = 'Client:      '
        ws['A1'].value += '    '.join([client_name, superintendent_name])

        ws['A1'].font = Font(name='Lao UI', size=10, bold=True)
        ws['A2'].value = '             '
        ws['A2'].value += '    '.join([client_first_line_address, superintendent_first_line_address])
        ws['A3'].value = '             '
        ws['A3'].value += '    '.join([client_second_line_address, superintendent_second_line_address])
        ws['C1'].value = f'Reference #: {arrow.now("Australia/Canberra").format("MM")}-{self.reference_number}'
        ws['C1'].font = Font(name='Lao UI', size=10, bold=True)
        ws['C3'].value = f'Date: {arrow.now("Australia/Canberra").format("DD/MM/YY")}'
        ws['C3'].font = Font(name='Lao UI', size=10)

        ws['A4'].value = 'PROGRESS CLAIM No.'
        ws['A4'].font = Font(name='Lao UI', size=14)
        ws['A4'].fill = fill
        ws['B4'].value = f'Project No: {self.reference_number}'
        ws['B4'].font = Font(name='Lao UI', size=10)
        ws['C4'].value = f'{arrow.now("Australia/Canberra").format("MMMM")}'
        ws['C4'].font = Font(name='Lao UI', size=10)
        ws['B5'].value = 'Approval terms: '
        ws['B5'].font = Font(name='Lao UI', size=10)

        ws['A6'].value = f'Project: {self.name}'
        ws['A6'].font = Font(name='Lao UI', size=11, bold=True)
        ws['B6'].value = 'Contract'
        ws['B6'].font = Font(name='Lao UI', size=10, bold=True)
        ws['B6'].alignment = Alignment(horizontal='center')
        ws['C6'].value = 'Completed'
        ws['C6'].font = Font(name='Lao UI', size=10, bold=True)
        ws['C6'].alignment = Alignment(horizontal='center')
        ws['A7'].value = ''
        ws['B7'].value = 'Value'
        ws['B7'].font = Font(name='Lao UI', size=10, bold=True)
        ws['B7'].alignment = Alignment(horizontal='center')
        ws['C7'].value = 'To Date'
        ws['C7'].font = Font(name='Lao UI', size=10, bold=True)
        ws['C7'].alignment = Alignment(horizontal='center')
        ws.merge_cells('D6:D7')
        ws['D6'].value = '%'
        ws['D6'].font = Font(name='Arial', size=12, bold=True)
        ws['D6'].alignment = Alignment(horizontal='center', vertical='center')

        for row in [6, 7]:
            for column in 'ABCD':
                cell = ws[f'{column}{row}']
                cell.fill = fill

        ws['A1'].border = Border(top=Side(border_style='medium', color='FF000000'))
        ws['B1'].border = Border(
                top=Side(border_style='medium', color='FF000000'),
                right=Side(border_style='medium', color='FF000000'),
        )
        ws['C1'].border = Border(top=Side(border_style='medium', color='FF000000'))
        ws['D1'].border = Border(top=Side(border_style='medium', color='FF000000'))

        ws['B2'].border = Border(right=Side(border_style='medium', color='FF000000'))

        ws['A3'].border = Border(bottom=Side(border_style='medium', color='FF000000'))
        ws['B3'].border = Border(
                bottom=Side(border_style='medium', color='FF000000'),
                right=Side(border_style='medium', color='FF000000'),
        )

        ws['A4'].border = Border(
                top=Side(border_style='medium', color='FF000000'),
                bottom=Side(border_style='medium', color='FF000000'),
        )
        ws['A6'].border = Border(
                top=Side(border_style='medium', color='FF000000'),
                bottom=Side(border_style='medium', color='FF000000'),
                left=Side(border_style='medium', color='FF000000'),
                right=Side(border_style='medium', color='FF000000'),
        )
        ws['B6'].border = Border(
                top=Side(border_style='medium', color='FF000000'),
                left=Side(border_style='medium', color='FF000000'),
                right=Side(border_style='medium', color='FF000000'),
        )
        ws['C6'].border = Border(
                top=Side(border_style='medium', color='FF000000'),
                left=Side(border_style='medium', color='FF000000'),
                right=Side(border_style='medium', color='FF000000'),
        )
        ws['D6'].border = Border(
                top=Side(border_style='medium', color='FF000000'),
                left=Side(border_style='medium', color='FF000000'),
                right=Side(border_style='medium', color='FF000000'),
        )
        ws['A7'].border = Border(
                top=Side(border_style='medium', color='FF000000'),
                bottom=Side(border_style='medium', color='FF000000'),
                left=Side(border_style='medium', color='FF000000'),
                right=Side(border_style='medium', color='FF000000'),
        )
        ws['B7'].border = Border(
                bottom=Side(border_style='medium', color='FF000000'),
                left=Side(border_style='medium', color='FF000000'),
                right=Side(border_style='medium', color='FF000000'),
        )
        ws['C7'].border = Border(
                bottom=Side(border_style='medium', color='FF000000'),
                left=Side(border_style='medium', color='FF000000'),
                right=Side(border_style='medium', color='FF000000'),
        )
        ws['D7'].border = Border(
                bottom=Side(border_style='medium', color='FF000000'),
                left=Side(border_style='medium', color='FF000000'),
                right=Side(border_style='medium', color='FF000000'),
        )

        row = 8

        def _comp(p):
            return p.id

        self.progress_items.sort(key=_comp)
        for progress_item in self.progress_items:
            ws[f'A{row}'].value = progress_item.name
            ws[f'B{row}'].value = progress_item.contract_value
            ws[f'C{row}'].value = progress_item.completed_value
            ws[f'D{row}'].value = f'= C{row}/B{row}'
            # print(ws['B{}'.format(row)].number_format)
            # print(ws['C{}'.format(row)].number_format)
            row += 1

        for irow in range(8, row):
            for column in 'ABCD':
                cell = ws[f'{column}{irow}']
                cell.font = Font(name='Lao UI', size=9)
                cell.border = Border(
                        left=Side(border_style='thin', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000'),
                )
                if column == 'D':
                    cell.alignment = Alignment(vertical='center', horizontal='center')

        ws[f'A{row}'].value = 'TOTAL OF CONTRACT'
        ws[f'B{row}'].value = f'=SUM(B{8}:B{row - 1})'
        ws[f'C{row}'].value = f'=SUM(C{8}:C{row - 1})'
        ws[f'D{row}'].value = f'=C{row}/B{row}'

        for column in 'ABCD':
            cell = ws[f'{column}{row}']
            cell.font = Font(name='Lao UI', size=9, bold=True)

        row += 1

        ws[f'A{row}'].value = 'Variations - See Appendix A attached'
        ws[f'B{row}'].value = r"='Appendix A'!D34"
        ws[f'C{row}'].value = r"='Appendix A'!E34"
        ws[f'D{row}'].value = f'=C{row}/B{row}'
        for row in [row - 1, row]:
            for column in 'ABCD':
                cell = ws[f'{column}{row}']
                cell.fill = fill
                cell.border = Border(
                        top=Side(border_style='thin', color='FF000000'),
                        bottom=Side(border_style='thin', color='FF000000'),
                        left=Side(border_style='thin', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000')
                )

        row += 1
        ws[f'A{row}'].value = 'Totals Excluding GST'
        ws[f'B{row}'].value = f'=B{row - 1} + B{row - 2}'
        ws[f'C{row}'].value = f'=C{row - 1} + C{row - 2}'
        ws[f'D{row}'].value = f'=C{row}/B{row}'
        for column in 'BCD':
            cell = ws[f'{column}{row}']
            cell.fill = fill
            cell.border = Border(
                    top=Side(border_style='medium', color='FF000000'),
                    bottom=Side(border_style='medium', color='FF000000'),
                    left=Side(border_style='medium', color='FF000000'),
                    right=Side(border_style='medium', color='FF000000')
            )

        row += 1
        ws[f'A{row}'].value = 'Less paid to date'
        ws[f'C{row}'].border = Border(
                left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000')
        )

        row += 1
        ws[f'A{row}'].value = 'Value of work completed this period'
        ws[f'C{row}'].value = f'=C{row - 2} - C{row - 1}'
        ws[f'C{row}'].border = Border(
                left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000')
        )

        row += 1
        ws[f'A{row}'].value = 'GST this period'
        ws[f'C{row}'].value = f'=C{row - 1} * 10%'
        ws[f'C{row}'].border = Border(
                left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000')
        )
        for irow in range(row - 3, row + 1):
            for column in 'ABCD':
                cell = ws[f'{column}{irow}']
                cell.font = Font(name='Lao UI', size=9)

        row += 1
        ws[f'A{row}'].value = 'TOTAL PAYABLE THIS CLAIM'
        ws[f'A{row}'].font = Font(name='Lao UI', size=9, bold=True)
        ws[f'C{row}'].value = f'=C{row - 2} + C{row - 1}'
        ws[f'C{row}'].font = Font(name='Lao UI', size=9, bold=True)
        ws[f'C{row}'].border = Border(
                top=Side(border_style='medium', color='FF000000'),
                bottom=Side(border_style='medium', color='FF000000'),
                left=Side(border_style='medium', color='FF000000'),
                right=Side(border_style='medium', color='FF000000')
        )

        for irow in range(8, row + 1):
            ws[f'B{irow}'].number_format = r'_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'
            ws[f'C{irow}'].number_format = r'_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'
            ws[f'D{irow}'].number_format = r'0.00%'

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 8
        ws.sheet_view.view = 'pageLayout'

        ws: openpyxl.worksheet.Worksheet = wb.create_sheet()
        ws.title = 'Appendix A'
        prepare(ws)

        ws.merge_cells('A1:D1')
        ws['A1'].style.alignment.wrap_text = True
        ws['A1'].value = f'{self.name}\nJOB #: {self.reference_number}'
        ws['A1'].fill = fill
        ws['A1'].font = Font(name='Lao UI', size=11, bold=True)

        ws.merge_cells('A3:B3')
        ws['A3'].value = "Appendix 'A' - Contract variations"
        ws['A3'].font = Font(name='Lao UI', size=10, bold=True)
        # ws['E3'].value = '=TODAY()'
        ws['E3'].value = datetime.today()
        ws['E3'].number_format = 'mm-dd-yy'
        ws['E3'].font = Font(name='Lao UI', size=9)

        ws['A5'].value = 'NO.'
        ws['B5'].value = 'ITEM'
        ws['C5'].value = 'PENDING'
        ws['D5'].value = 'APPROVED'
        ws['E5'].value = 'COMPLETED'
        for cell in ['A5', 'B5', 'C5', 'D5', 'E5']:
            ws[cell].font = Font(name='Lao UI', size=10, bold=True)
            ws[cell].alignment = Alignment(horizontal='center')
            ws[cell].fill = fill
            ws[cell].border = Border(
                    top=Side(border_style='medium', color='FF000000'),
                    bottom=Side(border_style='medium', color='FF000000'),
                    left=Side(border_style='medium', color='FF000000'),
                    right=Side(border_style='medium', color='FF000000')
            )

        row = 6
        self.variations.sort(key=lambda v: v.vid)
        for variation in self.variations:
            ws['A' + str(row)].value = row - 5
            ws['B' + str(row)].value = variation.description

            if variation.pending or variation.approved:
                if variation.pending:
                    column = 'C'
                else:
                    column = 'D'
                ws[column + str(row)].value = variation.amount
            elif variation.declined:
                ws['B' + str(row)].value = f"{variation.description} (declined {variation.amount})"
                ws['C' + str(row)].value = 0.0

            if variation.completed:
                ws['E' + str(row)].value = variation.amount

            row += 1

        while row < 34:
            for column in ['A', 'B', 'C', 'D', 'E']:
                cell = ws[column + str(row)]
                cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                     right=Side(border_style='thin', color='FF000000'))
            row += 1

        for index in range(6, row):
            for column in ['A', 'B', 'C', 'D', 'E']:
                cell = ws[f'{column}{index}']

                cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                     right=Side(border_style='thin', color='FF000000'))

                if column == 'A':
                    cell.font = Font(name='Lao UI', size=10, bold=True)
                    cell.alignment = Alignment(vertical='center', horizontal='center')
                else:
                    cell.font = Font(name='Lao UI', size=9)
                    if column == 'B':
                        cell.alignment = Alignment(vertical='center', horizontal='left')
                    else:
                        cell.alignment = Alignment(vertical='center')
                        cell.number_format = r'_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'

        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].value = 'TOTALS'
        ws[f'C{row}'].value = f'=SUM(C6:C{row - 1})'
        ws[f'D{row}'].value = f'=SUM(D6:D{row - 1})'
        ws[f'E{row}'].value = f'=SUM(E6:E{row - 1})'

        for column in 'ABCDE':
            cell = ws[f'{column}{row}']
            cell.alignment = Alignment(vertical='center', horizontal='center')
            cell.fill = fill
            cell.border = Border(
                    left=Side(border_style='medium', color='FF000000'),
                    right=Side(border_style='medium', color='FF000000'),
                    top=Side(border_style='medium', color='FF000000'),
                    bottom=Side(border_style='medium', color='FF000000')
            )
            if column != 'A':
                cell.number_format = r'_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'

        ws.row_dimensions[1].height = 30
        for row in range(len(self.variations)):
            ws.row_dimensions[6 + row].height = 30
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 13

        ws.cell('A1').alignment = Alignment(wrapText=True)
        ws.sheet_view.view = 'pageLayout'

        for index, variation in enumerate(self.variations):
            new_ws: openpyxl.worksheet.Worksheet = wb.create_sheet()
            prepare(new_ws)
            new_ws.title = f'V{index + 1}'

            new_ws.merge_cells('A1:H1')
            new_ws['A1'].value = 'CONTRACT VARIATION'
            new_ws['A1'].fill = fill
            new_ws['A1'].font = Font(name='Lao UI', size=16, bold=True)
            new_ws['A1'].alignment = Alignment(vertical='center', horizontal='center')

            for column in 'ABCDEFGH':
                new_ws[column + '1'].border = Border(top=Side(border_style='medium', color='FF000000'),
                                                     bottom=Side(border_style='medium', color='FF000000'),
                                                     left=Side(border_style='medium', color='FF000000'),
                                                     right=Side(border_style='medium', color='FF000000'))

            new_ws.merge_cells('B3:H3')
            new_ws['B3'].value = f'PROJECT: {self.name}'
            new_ws['B3'].fill = fill
            new_ws['B3'].font = Font(name='Lao UI', size=12, bold=True)
            for column in 'BCDEFGH':
                new_ws[f'{column}3'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                     bottom=Side(border_style='thin', color='FF000000'))

            new_ws['B5'].value = 'VARIATION NO:'
            new_ws['B5'].fill = fill
            new_ws['C5'].fill = fill
            new_ws['B5'].font = Font(name='Lao UI', size=12, bold=True)
            new_ws.merge_cells('D5:E5')
            new_ws['D5'].value = index + 1
            new_ws['D5'].fill = fill
            new_ws['D5'].font = Font(name='Lao UI', size=14, bold=True)
            new_ws.merge_cells('G5:H5')
            new_ws['G5'].value = f'TPC REF: {self.reference_number}'
            new_ws['G5'].font = Font(name='Lao UI', size=14, bold=True)
            new_ws['G5'].alignment = Alignment(vertical='center', horizontal='left')
            for column in 'BCDE':
                cell = new_ws[column + str(5)]
                cell.border = Border(top=Side(border_style='thin', color='FF000000'),
                                     bottom=Side(border_style='thin', color='FF000000'))

            for column in 'BCDEFGH':
                cell = new_ws[column + str(6)]
                cell.border = Border(bottom=Side(border_style='medium', color='FF000000'))

            row = 7
            if len(variation.items) > 1:
                new_ws.merge_cells(f'B{row}:G{row}')
                new_ws[f'B{row}'].value = variation.description
                new_ws[f'B{row}'].font = Font(name='Lao UI', size=11, bold=True)
                new_ws[f'B{row}'].alignment = Alignment(vertical='center')
                row += 1

            for item in variation.items:
                new_ws.merge_cells(f'B{row}:G{row}')
                new_ws[f'B{row}'].value = item.description
                new_ws[f'B{row}'].font = Font(name='Lao UI', size=11, bold=True)
                new_ws[f'B{row}'].alignment = Alignment(vertical='center')

                new_ws[f'H{row}'].value = item.amount
                new_ws[f'H{row}'].font = Font(name='Lao UI', size=11, bold=True)
                new_ws[f'H{row}'].alignment = Alignment(vertical='center')
                new_ws[f'H{row}'].number_format = r'_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'

                new_ws.row_dimensions[row].height = 40

                row += 1

            while row < 13:
                new_ws[f'B{row}'].font = Font(name='Lao UI', size=11)
                new_ws[f'H{row}'].font = Font(name='Lao UI', size=11)
                row += 1

            for column in 'BCDEFGH':
                cell = new_ws[column + str(row - 1)]
                cell.border = Border(bottom=Side(border_style='thin', color='FF000000'))

            new_ws['B' + str(row)].value = 'Value of work'
            new_ws['H' + str(row)].value = f'=SUM(H7:H{row - 1})'
            new_ws['B' + str(row)].font = Font(name='Lao UI', size=11)
            new_ws['H' + str(row)].font = Font(name='Lao UI', size=11)
            row += 1

            new_ws['B' + str(row)].value = f'Add OH/Profit {self.margin * 100}%'
            new_ws['H' + str(row)].value = f'=H{row - 1} * {self.margin * 100}%'
            new_ws['B' + str(row)].font = Font(name='Lao UI', size=11)
            new_ws['H' + str(row)].font = Font(name='Lao UI', size=11)

            if self.admin_fee is not None and self.admin_fee != 0:
                row += 1
                new_ws['B' + str(row)].value = 'Fixed administration fee'
                new_ws['H' + str(row)].value = self.admin_fee
                new_ws['B' + str(row)].font = Font(name='Lao UI', size=11)
                new_ws['H' + str(row)].font = Font(name='Lao UI', size=11)

            for column in 'BCDEFGH':
                new_ws[column + str(row)].border = Border(bottom=Side(border_style='thin', color='FF000000'))
            row += 1

            new_ws['B' + str(row)].value = 'Subtotal'
            if self.admin_fee is None:
                new_ws['H' + str(row)] = f'=H{row - 2} + H{row - 1}'
            else:
                new_ws['H' + str(row)] = f'=H{row - 3} + H{row - 2} + H{row - 1}'
            new_ws['H' + str(row)].font = Font(name='Lao UI', size=11, bold=True)
            row += 1

            new_ws['B' + str(row)] = 'Add GST'
            new_ws['H' + str(row)].value = f'=H{row - 1} * 0.1'
            new_ws['H' + str(row)].font = Font(name='Lao UI', size=11, underline='singleAccounting')
            for column in 'BCDEFGH':
                new_ws[column + str(row)].border = Border(bottom=Side(border_style='medium', color='FF000000'))
            row += 1

            new_ws.merge_cells(f'B{row}:C{row}')
            for column in 'BCDEFGH':
                new_ws[f'{column}{row}'].fill = fill
            new_ws['B' + str(row)] = 'TOTAL'
            new_ws['H' + str(row)].value = f'=H{row - 1} + H{row - 2}'
            new_ws['H' + str(row)].font = Font(name='Lao UI', size=11, bold=True)

            for idx in range(7, row + 1):
                new_ws['H' + str(idx)].number_format = r'_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'

            row += 4
            new_ws[f'B{row}'].value = 'Variation Prepared By:'
            new_ws[f'G{row}'].value = 'Variation Prepared For:'
            new_ws[f'B{row}'].font = Font(name='Lao UI', size=11)
            new_ws[f'G{row}'].font = Font(name='Lao UI', size=11)

            row += 5
            new_ws[f'B{row}'].value = 'FOR'
            new_ws[f'B{row}'].font = Font(name='Lao UI', size=11)

            new_ws[f'B{row + 1}'].value = 'Total Project Construction Pty Ltd'
            new_ws[f'B{row + 1}'].font = Font(name='Lao UI', size=11)

            new_ws[f'G{row}'].value = 'FOR'
            new_ws[f'G{row}'].font = Font(name='Lao UI', size=11)

            new_ws[f'G{row + 1}'].value = variation.prepared_for
            new_ws[f'G{row + 1}'].font = Font(name='Lao UI', size=11)

            row += 3
            new_ws[f'B{row}'].value = 'Date:'
            # new_ws['C{}'.format(row)].value = '=TODAY()'
            new_ws[f'C{row}'].value = datetime.today()
            new_ws[f'C{row}'].number_format = 'mm-dd-yy'
            new_ws[f'G{row}'].value = 'Date:'
            new_ws[f'B{row}'].font = Font(name='Lao UI', size=11)
            new_ws[f'C{row}'].font = Font(name='Lao UI', size=11)
            new_ws[f'G{row}'].font = Font(name='Lao UI', size=11)

            new_ws.row_dimensions[1].height = 60
            new_ws.row_dimensions[3].height = 40
            new_ws.column_dimensions['A'].width = 3
            new_ws.column_dimensions['D'].width = 6
            new_ws.column_dimensions['F'].width = 4
            new_ws.column_dimensions['H'].width = 12
            new_ws.sheet_view.view = 'pageLayout'

        fn = self.name + '.xlsx'
        wb.save('generated/' + fn)
        return fn


class Variation(db.Model):
    __tablename__ = 'variations'
    vid: int = db.Column(db.Integer, primary_key=True)
    date: datetime = db.Column(db.DateTime, default=datetime.utcnow())
    subcontractor: str = db.Column(db.String(64), default="")
    invoice_no: str = db.Column(db.String(64), nullable=True)
    description: str = db.Column(db.Text, nullable=False)
    amount: float = db.Column(db.Float, default=0.0)
    pending: bool = db.Column(db.Boolean, default=True)
    approved: bool = db.Column(db.Boolean, default=False)
    declined: bool = db.Column(db.Boolean, default=False)
    completed: bool = db.Column(db.Boolean, default=False)
    note: str = db.Column(db.Text, nullable=True)
    prepared_for: str = db.Column(db.Text, default='')
    project_id: int = db.Column(db.Integer, db.ForeignKey('projects.pid'))
    items: List[Item] = db.relationship('Item', backref='variation', cascade="all, delete-orphan", lazy='select')

    def __repr__(self) -> str:
        return f'<Variation Id: {self.vid}, Project id: {self.project_id}, Description: {self.description}, Amount: {self.amount}, Subcontractor: {self.subcontractor}>'

    def to_json(self) -> Dict[str, Any]:
        json_variation = {
            'vid': self.vid,
            'url': url_for('api.get_variation', variation_id=self.vid, _external=True),
            'date': self.date,
            'subcontractor': self.subcontractor,
            'invoice_no': self.invoice_no,
            'description': self.description,
            'amount': self.amount,
            'pending': self.pending,
            'approved': self.approved,
            'declined': self.declined,
            'completed': self.completed,
            'note': self.note,
            'prepared_for': self.prepared_for,
            'project_id': url_for('api.get_project', project_id=self.project_id, _external=True),
            'items': url_for('api.get_variation_items', variation_id=self.vid, _external=True)
        }
        return json_variation

    @staticmethod
    def from_json(json_variation: Dict[str, Any]) -> 'Variation':
        date = arrow.get(json_variation.get('date')).datetime
        subcontractor = json_variation.get('subcontractor')
        invoice_no = json_variation.get('invoice_no')
        description = json_variation.get('description')
        amount = json_variation.get('amount')
        pending = json_variation.get('pending', True)
        approved = json_variation.get('approved', False)
        declined = json_variation.get('declined', False)
        completed = json_variation.get('completed', False)
        note = json_variation.get('note', '')
        prepared_for = json_variation.get('prepared_for', '')
        project = Project.query.get_or_404(int(json_variation.get('project_id')))

        return Variation(date=date, subcontractor=subcontractor, invoice_no=invoice_no, description=description,
                         amount=amount, pending=pending, approved=approved, declined=declined, completed=completed,
                         note=note, project=project, prepared_for=prepared_for)


class Item(db.Model):
    __tablename__ = 'items'
    id: int = db.Column(db.Integer, primary_key=True)
    amount: float = db.Column(db.Float, nullable=False)
    description: str = db.Column(db.Text, nullable=False)
    variation_id: int = db.Column(db.Integer, db.ForeignKey('variations.vid'))

    def __repr__(self) -> str:
        return f'<Item Id: {self.id}, Variation: {self.variation_id}, Amount: {self.amount}, Description: {self.description}>'

    def to_json(self) -> Dict[str, Any]:
        return {
            "id": self.id,
            "variation_id": self.variation_id,
            "amount": self.amount,
            "description": self.description,
            'url': url_for('api.get_item', item_id=self.id, _external=True)
        }

    @staticmethod
    def from_json(json: Dict[str, Any]) -> 'Item':
        variation = Variation.query.get_or_404(int(json.get('variation_id')))
        description = json.get('description')
        amount = json.get('amount')
        return Item(variation=variation, amount=amount, description=description)


class ProgressItem(db.Model):
    __tablename__ = 'progress_items'
    id: int = db.Column(db.Integer, primary_key=True)
    name: str = db.Column(db.String, nullable=False)
    contract_value: float = db.Column(db.Float, nullable=False)
    completed_value: float = db.Column(db.Float, default=0.0)

    project_id: int = db.Column(db.Integer, db.ForeignKey('projects.pid'), nullable=False)

    def to_json(self) -> Dict[str, Any]:
        if self.contract_value != 0:
            percentage = self.completed_value / self.contract_value
        else:
            percentage = 0
        return {
            'id': self.id,
            'name': self.name,
            'contract_value': self.contract_value,
            'completed_value': self.completed_value,
            'project_id': self.project_id,
            'percentage': percentage
        }

    @staticmethod
    def from_json(json: Dict[str, Any]) -> 'ProgressItem':
        project: Project = Project.query.get_or_404(int(json.get('project_id')))
        name = json['name']
        contract_value = json['contract_value']
        completed_value = None
        if 'completed_value' in json:
            completed_value = json['completed_value']
        return ProgressItem(name=name, contract_value=contract_value, completed_value=completed_value, project=project)


class Client(db.Model):
    __tablename__ = 'clients'
    id: int = db.Column(db.Integer, primary_key=True)
    name: str = db.Column(db.String, nullable=False)
    first_line_address: str = db.Column(db.String, nullable=True)
    second_line_address: str = db.Column(db.String, nullable=True)
    project_id: int = db.Column(db.Integer, db.ForeignKey('projects.pid'), nullable=False)

    def to_json(self) -> Dict[str, Any]:
        return {
            'id': self.id,
            'name': self.name,
            'first_line_address': self.first_line_address,
            'second_line_address': self.second_line_address
        }

    @staticmethod
    def from_json(json: Dict[str, Any]) -> 'Client':
        project: Project = Project.query.get_or_404(int(json.get('project_id')))
        name = json['name']
        first_line_address = None
        if 'first_line_address' in json:
            first_line_address = json['first_line_address']
        second_line_address = None
        if 'second_line_address' in json:
            second_line_address = json['second_line_address']
        return Client(name=name, project=project, first_line_address=first_line_address,
                      second_line_address=second_line_address)


class Superintendent(db.Model):
    __tablename__ = 'superintendents'
    id: int = db.Column(db.Integer, primary_key=True)
    name: str = db.Column(db.String, nullable=False)
    first_line_address: str = db.Column(db.String, nullable=True)
    second_line_address: str = db.Column(db.String, nullable=True)
    project_id: int = db.Column(db.Integer, db.ForeignKey('projects.pid'), nullable=False)

    def to_json(self) -> Dict[str, Any]:
        return {
            'id': self.id,
            'name': self.name,
            'first_line_address': self.first_line_address,
            'second_line_address': self.second_line_address
        }

    @staticmethod
    def from_json(json: Dict[str, Any]) -> 'Superintendent':
        project: Project = Project.query.get_or_404(int(json.get('project_id')))
        name = json['name']
        first_line_address = None
        if 'first_line_address' in json:
            first_line_address = json['first_line_address']
        second_line_address = None
        if 'second_line_address' in json:
            second_line_address = json['second_line_address']
        return Superintendent(name=name, project=project, first_line_address=first_line_address,
                              second_line_address=second_line_address)
